# custom_plotter.py

import os
import pandas as pd
import numpy as np
from scipy.integrate import simpson
from openpyxl import load_workbook
import xlrd
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D

def read_xls(file_path):
    workbook = xlrd.open_workbook(file_path)
    sheet_names = workbook.sheet_names()
    return workbook, sheet_names

def read_xlsx(file_path):
    workbook = load_workbook(file_path)
    sheet_names = workbook.sheetnames
    return workbook, sheet_names

def make_plot(file_path, ax, color, linestyle):
    tensile = np.array([])
    elon    = np.array([])

    if file_path.endswith('.xlsx'):
        workbook, sheet_names = read_xlsx(file_path)
    elif file_path.endswith('.xls'):
        workbook, sheet_names = read_xls(file_path)
    else:
        raise ValueError("Unsupported file format")

    toughness = []
    for i, sheet_name in enumerate(sheet_names):
        if i == 0 or i == 1:
            continue

        if file_path.endswith('.xlsx'):
            sheet = workbook[sheet_name]
            data = pd.DataFrame(sheet.values)
            data.columns = data.iloc[1]
            data = data.drop(0)
        elif file_path.endswith('.xls'):
            sheet = workbook.sheet_by_name(sheet_name)
            data = []
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                data.append(row)
            data = pd.DataFrame(data)
            data.columns = data.iloc[1]
            data = data.drop([0, 1])  # Adjusting to drop initial non-data rows

            try:
                elongation = data['Strain'][4:].astype(float)
                stress = data['Standard force'][4:].astype(float)
            except:
                elongation = data['Dehnung'][4:].astype(float)
                stress = data['Standardkraft'][4:].astype(float)

        area = simpson(stress, elongation)
        tensile = np.append(tensile, stress)
        elon    = np.append(elon, elongation)

        toughness.append(area)
        ax.plot(elongation, stress, linestyle=linestyle, color=color)

def plot_custom_files(files):
    fig, ax = plt.subplots(figsize=(6, 6))

    colors = plt.cm.tab20(np.linspace(0, 1, len(files)))
    linestyles = [":", "--", "-", "-."] * (len(files) // 4 + 1)

    legend_elements = []

    for file, color, linestyle in zip(files, colors, linestyles):
        make_plot(file, ax, color, linestyle)
        legend_elements.append(Line2D([0], [0], color=color, lw=2, linestyle=linestyle, label=os.path.splitext(os.path.basename(file))[0]))

    ax.set_xlabel("Elongation [%]")
    ax.set_ylabel("Tensile Strength [MPa]")

    plt.grid(color='gray', linestyle=':', linewidth=0.5)
    plt.legend(handles=legend_elements, loc='lower right')

    plt.tight_layout()
    return fig, ax

